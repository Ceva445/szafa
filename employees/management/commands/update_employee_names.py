from django.core.management.base import BaseCommand
from employees.models import Employee
from openpyxl import load_workbook


class Command(BaseCommand):
    help = "Update employee names from Excel file"

    def add_arguments(self, parser):
        parser.add_argument("file", type=str, help="Path to Excel file")

    def handle(self, *args, **options):
        file_path = options["file"]

        wb = load_workbook(file_path)
        sheet = wb["Sheet1"]

        updated = 0
        not_found = 0

        for row in sheet.iter_rows(min_row=2, values_only=True):
            card_number, last_name, first_name = row

            card_number = str(card_number).strip()

            try:
                employee = Employee.objects.get(card_number=card_number)

                employee.first_name = first_name.strip()
                employee.last_name = last_name.strip()
                employee.save(update_fields=["first_name", "last_name"])

                updated += 1

            except Employee.DoesNotExist:
                self.stdout.write(
                    self.style.WARNING(f"Employee not found: {card_number}")
                )
                not_found += 1

        self.stdout.write(
            self.style.SUCCESS(
                f"Updated: {updated}, Not found: {not_found}"
            )
        )